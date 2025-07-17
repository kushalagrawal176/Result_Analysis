import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

def TE_analysis(file1, file2, sub):
    # Load the existing workbook
    workbook = load_workbook(file1)
    sheet = workbook["Sheet1"]

    # Check if the "Result Analysis" sheet exists
    if 'Result Analysis' in workbook.sheetnames:
        result_sheet = workbook['Result Analysis']
    else:
        result_sheet = workbook.create_sheet('Result Analysis')

    # Clear existing content in the "Result Analysis" sheet if it exists
    for row in result_sheet.iter_rows(min_row=1, max_row=result_sheet.max_row, min_col=1, max_col=result_sheet.max_column):
        for cell in row:
            cell.value = None

    # Define the row and column headings
    row_headings = ["Subjects", "Appeared", "Pass", "Fail", "Absent", "Dist", "Dist %", "FC", "FC %", "HSC", "HSC %", "SC", "SC %", "PC", "PC %", "Fail", "Fail %", "Total Count", "Total %", "Sub./Class Result"]

    for i in range(0, len(sub)):
        result_sheet.cell(row = 1, column=i+2, value = sub[i][1])
    result_sheet.cell(row = 1, column = len(sub)+2, value = "SGPA")

    # Set the row headings in the first column of the result sheet
    for i, heading in enumerate(row_headings, start=1):
        result_sheet.cell(row=i, column=1, value=heading)

    # Function to calculate numeric values from text (first three characters) for CGPA
    def get_cgpa_value(cell_value):
        try:
            return float(str(cell_value).strip()[:4])  # Extract CGPA as float, considering up to 4 digits
        except:
            return None

    # Helper function to calculate percentages
    def calculate_percentage(part, whole):
        return round(part/whole * 100, 2)

    def get_numeric_value(cell_value):
        """
        Function to trim whitespaces, extract the first 3 characters of the string,
        convert it to a numeric value, and return the result.
        """
        try:
            # Trim whitespaces and extract the first 3 characters
            trimmed_value = str(cell_value).strip()[:3]
            # Convert the extracted value to a numeric value
            return float(trimmed_value)
        except (ValueError, TypeError):
            # Return None if conversion fails or the value is invalid
            return None

    sgpa_1 = []
    final_arr_1 = []

    # Loop through each non-null column and calculate results
    for col_idx in range(5, sheet.max_column+1):
        values = [sheet.cell(row=row, column=col_idx).value for row in range(4, sheet.max_row + 1)]

        if all(v is None for v in values):
            continue
        
        if col_idx < sheet.max_column:
            numeric_values = [get_numeric_value(v) for v in values if v is not None and v != 'AB']

            numeric_values = [v for v in numeric_values if isinstance(v, (int, float))]

            appeared = len([v for v in numeric_values if v >= 0])
            passed = sum(1 for v in numeric_values if v >= 40)
            failed = sum(1 for v in numeric_values if v < 40)
            absent = values.count('AB')
            dist = sum(1 for v in numeric_values if v >= 66)
            fc = sum(1 for v in numeric_values if v >= 60) - dist
            hsc = sum(1 for v in numeric_values if v >= 55) - fc - dist
            sc = sum(1 for v in numeric_values if v >= 50) - hsc - fc - dist
            pc = sum(1 for v in numeric_values if v >= 40) - sc - hsc - fc - dist
        
            dist_perc = calculate_percentage(dist, appeared)
            fc_perc = calculate_percentage(fc, appeared)
            hsc_perc = calculate_percentage(hsc, appeared)
            sc_perc = calculate_percentage(sc, appeared)
            pc_perc = calculate_percentage(pc, appeared)
            fail_perc = calculate_percentage(failed, appeared)
            subject_class_result = calculate_percentage(passed, appeared)
        
            total_count = dist + fc + hsc + sc + pc + failed
            total_perc = calculate_percentage(total_count, appeared)
        
            results = [
                appeared, passed, failed, absent, dist, dist_perc, fc, fc_perc, hsc, hsc_perc, sc, sc_perc, pc, pc_perc, failed, fail_perc, total_count, total_perc, subject_class_result
            ]
        
            for i, result in enumerate(results, start=2):
                result_sheet.cell(row=i, column=col_idx-3, value=result)

            final_arr_1.append(round(passed/appeared * 100, 2))

        else:
            cgpa_values = [get_cgpa_value(v) for v in values if v not in [None, '--', 'AB']]  # Get valid CGPA values

            # Filter out any None values before proceeding with calculations
            cgpa_values = [v for v in cgpa_values if v is not None]

            passed = sum(1 for v in cgpa_values if v is not None and str(v).strip() != '--')  # Passed: not None, not '--'
            failed = sum(1 for v in values if str(v).strip() in [None, '--'])  # Failed: None or '--'
            appeared = passed + failed
            absent = values.count('AB')  # Absent: 'AB'
        
            dist = sum(1 for v in cgpa_values if v >= 7.75)  # Distinction
            fc = sum(1 for v in cgpa_values if 6.75 <= v < 7.75)  # First Class
            hsc = sum(1 for v in cgpa_values if 6.25 <= v < 6.75)  # Higher Second Class
            sc = sum(1 for v in cgpa_values if 5.5 <= v < 6.25)  # Second Class
            pc = sum(1 for v in cgpa_values if v < 5.5)  # Pass Class
        
            dist_perc = calculate_percentage(dist, appeared) if calculate_percentage(dist, appeared) <= 100.00 else 100.00
            fc_perc = calculate_percentage(fc, appeared) if calculate_percentage(fc, appeared) <= 100.00 else 100.00
            hsc_perc = calculate_percentage(hsc, appeared) if calculate_percentage(hsc, appeared) <= 100.00 else 100.00
            sc_perc = calculate_percentage(sc, appeared) if calculate_percentage(sc, appeared) <= 100.00 else 100.00
            pc_perc = calculate_percentage(pc, appeared) if calculate_percentage(pc, appeared) <= 100.00 else 100.00
            fail_perc = calculate_percentage(failed, appeared) if calculate_percentage(failed, appeared) <= 100.00 else 100.00
            subject_class_result = calculate_percentage(passed, appeared) if calculate_percentage(passed, appeared) <= 100.00 else 100.00
        
            total_count = dist + fc + hsc + sc + pc + failed
            total_perc = dist_perc + fc_perc + hsc_perc + sc_perc + pc_perc + fail_perc if dist_perc + fc_perc + hsc_perc + sc_perc + pc_perc + fail_perc <= 100.00 else 100.00
        
            # Results for the row
            results = [
                appeared, passed, failed, absent, dist, dist_perc, fc, fc_perc, hsc, hsc_perc, sc, sc_perc, pc, pc_perc, failed, fail_perc, total_count, total_perc, subject_class_result
            ]
        
            # Writing results in the result sheet for the current subject
            for i, result in enumerate(results, start=2):
                result_sheet.cell(row=i, column=col_idx-3, value=result)

            sgpa_1 = [dist, dist_perc, fc, fc_perc, hsc, hsc_perc, sc, sc_perc, pc, pc_perc, failed, fail_perc, passed, round(passed/appeared * 100, 2)]

    def create_failed_students_sheet(workbook):
        # Access the source sheet
        sheet = workbook['Sheet1']

        # Check if the "Failed Students" sheet exists, otherwise create it
        if 'Failed Students' in workbook.sheetnames:
            failed_sheet = workbook['Failed Students']
        else:
            failed_sheet = workbook.create_sheet('Failed Students')

        failed_sheet.cell(row = 1, column = 1, value = "Sr. No.")
        failed_sheet.cell(row = 1, column = 2, value = "Roll No.")
        failed_sheet.cell(row = 1, column = 3, value = "Seat No.")
        failed_sheet.cell(row = 1, column = 4, value = "Name")

        for i in range(0, len(sub)):
            failed_sheet.cell(row = 1, column=i+5, value = sub[i][1])
        failed_sheet.cell(row = 1, column = len(sub)+5, value = "SGPA")

        # Start adding rows with failed students
        failed_row_index = 2  # Starting from the second row
        for row_idx in range(4, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column+1)]

            if(str(row_values[len(row_values)-1]).strip() == "--"):
                for col in range(1, len(row_values)+1):
                    failed_sheet.cell(row=failed_row_index, column=col, value=row_values[col-1])
                failed_row_index += 1


    create_failed_students_sheet(workbook)



    wb_prev = load_workbook(file2, data_only = True)
    sheet_prev = wb_prev['Result Analysis']
    last_col_prev = sheet_prev.max_column

    # Get sgpa_2 from rows 6 to 16, last column
    sgpa_2 = [float(sheet_prev.cell(row=i, column=last_col_prev).value) for i in range(6, 18)]
    sgpa_2.append(sheet_prev.cell(row=3, column=last_col_prev).value)
    sgpa_2.append(sheet_prev.cell(row=20, column=last_col_prev).value)

    # Get final_arr_2 from row 19, cols 2 to 13
    final_arr_2 = [float(sheet_prev.cell(row=20, column=j).value) for j in range(2, 16)]

    sheet_graphs = workbook.create_sheet('Graphs')

    sheet_graphs.append(["", "Curr Year", "Prev Year"])
    for i, label in enumerate(sub):
        row = [
            str(label[1]),  # Convert label to string, even if it's a tuple
            f"{final_arr_1[i]:.10f}"[:(f"{final_arr_1[i]:.10f}".find('.') + 3)],
            f"{final_arr_2[i]:.10f}"[:(f"{final_arr_2[i]:.10f}".find('.') + 3)]
        ]
        sheet_graphs.append(row)

    # Determine start row for second table
    second_table_start = len(sub) + 4
    sheet_graphs.cell(row=second_table_start, column=1).value = ""
    sheet_graphs.cell(row=second_table_start, column=2).value = "Curr Year"
    sheet_graphs.cell(row=second_table_start, column=3).value = "Prev Year"

    # Labels for 2nd table
    second_labels = ["Dist", "Dist %", "FC", "FC %", "HSC", "HSC %", "SC", "SC %", "PC", "PC %", "Fail", "Fail %", "Pass", "Pass %"]

    for i, label in enumerate(second_labels):
        val1 = f"{sgpa_1[i]:.10f}"[:(f"{sgpa_1[i]:.10f}".find('.') + 3)] if i < len(sgpa_1) else ""
        val2 = f"{sgpa_2[i]:.10f}"[:(f"{sgpa_2[i]:.10f}".find('.') + 3)] if i < len(sgpa_2) else ""
        row = [label, val1, val2]
        sheet_graphs.append(row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output