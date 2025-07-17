import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

def SE_analysis(file_path, prev_file, sub):
    wb = load_workbook(file_path)
    source_sheet_name = "Sheet1"
    result_sheet_name = "Result Analysis"
    fail_sheet_name = "Failed students"

    if source_sheet_name not in wb.sheetnames:
        print(f"Sheet '{source_sheet_name}' not found in the workbook.")
        return

    ws_source = wb[source_sheet_name]

    row_headings = ["Subjects", "Appeared", "Pass", "Fail", "Absent", "Dist", "Dist %", "FC", "FC %", 
                    "HSC", "HSC %", "SC", "SC %", "PC", "PC %", "Fail", "Fail %", "Total Count", "Total %", "Sub./Class Result"]

    if result_sheet_name in wb.sheetnames:
        ws_result = wb[result_sheet_name]
    else:
        ws_result = wb.create_sheet(result_sheet_name)

    for row_idx, row_name in enumerate(row_headings, start=1):
        ws_result.cell(row=row_idx, column=1, value=row_name)

    for i in range(0, len(sub)):
        ws_result.cell(row = 1, column=i+2, value = sub[i][1])
    ws_result.cell(row = 1, column = len(sub)+2, value = "SGPA")

    last_row = ws_source.max_row

    sgpa_1 = []
    final_arr_1 = []

    for col_idx in range(5, ws_source.max_column + 1):
        values = [ws_source.cell(row=row, column=col_idx).value for row in range(2, last_row + 1)]

        appeared = sum(1 for v in values if str(v).strip() not in [None, "--", "NA", "N/A", "AB"])

        if (col_idx == ws_source.max_column):
            thresholds = [7.75, 6.75, 6.25, 5.5, 5]
            pass_count = sum(1 for v in values if isinstance(v, (int, float)) and v >= 5)
            fail_count = sum(1 for v in values if (isinstance(v, (int, float)) and v < 5) or str(v).strip() == "--")
            appeared = pass_count + fail_count
        else:
            thresholds = [66, 60, 55, 50, 40]
            pass_count = sum(1 for v in values if isinstance(v, (int, float)) and v >= 40)
            fail_count = sum(1 for v in values if (isinstance(v, (int, float)) and v < 40))

        absent_count = sum(1 for v in values if (v == "AB" or str(v).strip() == "--"))

        dist, fc, hsc, sc, pc = [sum(1 for v in values if isinstance(v, (int, float)) and v >= t) for t in thresholds]
        fc -= dist
        hsc -= fc + dist
        sc -= hsc + fc + dist
        pc -= sc + hsc + fc + dist

        if(col_idx == ws_source.max_column):
            sgpa_1 = [dist, round(dist/appeared * 100, 2), fc, round(fc/appeared * 100, 2), hsc, round(hsc/appeared * 100, 2), sc, round(sc/appeared * 100, 2), pc, round(pc/appeared * 100, 2), fail_count, round(fail_count/appeared * 100, 2), pass_count, round(pass_count/appeared * 100, 2)]
        else:
            final_arr_1.append(round(pass_count/appeared * 100, 2))

        total_count = dist + fc + hsc + sc + pc + fail_count
        
        for row_idx, val in enumerate([appeared, pass_count, fail_count, absent_count, dist, 
                                       round((dist / appeared * 100), 2) if appeared else 0, fc, 
                                       round((fc / appeared * 100), 2) if appeared else 0, hsc, 
                                       round((hsc / appeared * 100), 2) if appeared else 0, sc, 
                                       round((sc / appeared * 100), 2) if appeared else 0, pc, 
                                       round((pc / appeared * 100), 2) if appeared else 0, fail_count,
                                       round((fail_count / appeared * 100), 2) if appeared else 0,
                                       total_count, sum([dist, fc, hsc, sc, pc, fail_count]) / appeared * 100 if appeared else 0,
                                       round((pass_count / appeared * 100), 2) if appeared else 0], start=2):
            ws_result.cell(row=row_idx, column=col_idx - 3, value=val)

    if fail_sheet_name in wb.sheetnames:
        ws_fail = wb[fail_sheet_name]
    else:
        ws_fail = wb.create_sheet(fail_sheet_name)

    ws_fail.cell(row = 1, column = 1, value = "Sr. No.")
    ws_fail.cell(row = 1, column = 2, value = "Roll No.")
    ws_fail.cell(row = 1, column = 3, value = "Seat No.")
    ws_fail.cell(row = 1, column = 4, value = "Name")

    for i in range(0, len(sub)):
        ws_fail.cell(row = 1, column=i+5, value = sub[i][1])
    ws_fail.cell(row = 1, column = len(sub)+5, value = "SGPA")

    last_row = ws_source.max_row
    fail_row = 2
    for row_idx in range(2, last_row + 1):
        row_values = [ws_source.cell(row=row_idx, column=col).value for col in range(1, ws_source.max_column + 1)]

        if(str(row_values[len(row_values)-1]).strip() == "--"):
            for col in range(1, len(row_values)+1):
                ws_fail.cell(row=fail_row, column=col, value=row_values[col-1])
            fail_row += 1


    wb_prev = load_workbook(prev_file, data_only = True)
    sheet_prev = wb_prev['Result Analysis']
    last_col_prev = sheet_prev.max_column

    # Get sgpa_2 from rows 6 to 16, last column
    sgpa_2 = [float(sheet_prev.cell(row=i, column=last_col_prev).value) for i in range(6, 18)]
    sgpa_2.append(sheet_prev.cell(row=3, column=last_col_prev).value)
    sgpa_2.append(sheet_prev.cell(row=20, column=last_col_prev).value)

    # Get final_arr_2 from row 19, cols 2 to 13
    final_arr_2 = [float(sheet_prev.cell(row=20, column=j).value) for j in range(2, 14)]

    sheet_graphs = wb.create_sheet('Graphs')

    sheet_graphs.append(["", "Curr Year", "Prev Year"])
    for i, label in enumerate(sub):
        row = [
            str(label[1]),  # Convert label to string, even if it's a tuple
            f"{final_arr_1[i]:.10f}"[:(f"{final_arr_1[i]:.10f}".find('.') + 3)],
            f"{final_arr_2[i]:.10f}"[:(f"{final_arr_2[i]:.10f}".find('.') + 3)]
        ]
        sheet_graphs.append(row)

    # Determine start row for second table
    second_table_start = len(sub) + 4
    sheet_graphs.cell(row=second_table_start, column=1).value = ""
    sheet_graphs.cell(row=second_table_start, column=2).value = "Curr Year"
    sheet_graphs.cell(row=second_table_start, column=3).value = "Prev Year"

    # Labels for 2nd table
    second_labels = ["Dist", "Dist %", "FC", "FC %", "HSC", "HSC %", "SC", "SC %", "PC", "PC %", "Fail", "Fail %", "Pass", "Pass %"]

    for i, label in enumerate(second_labels):
        val1 = f"{sgpa_1[i]:.10f}"[:(f"{sgpa_1[i]:.10f}".find('.') + 3)] if i < len(sgpa_1) else ""
        val2 = f"{sgpa_2[i]:.10f}"[:(f"{sgpa_2[i]:.10f}".find('.') + 3)] if i < len(sgpa_2) else ""
        row = [label, val1, val2]
        sheet_graphs.append(row)


    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output